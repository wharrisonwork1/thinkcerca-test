import os
import re
import json
import pandas as pd
import fitz  # PyMuPDF
from tqdm import tqdm
from openai import OpenAI
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# --- ThinkCERCA imports ---
from thinkcerca_tool.config import (
    FILES,
    OUTPUT_FILE,
    MODEL_NAME,
    OPENAI_API_KEY,
    DATA_DIR,
    TARGET_MODULE,
)
from thinkcerca_tool.modules.join_standards import join_module_standards

# --- Output directory ---
OUTPUT_DIR = Path(DATA_DIR).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
#  PDF ACTIVITY EXTRACTION
# ============================================================
def extract_pdf_activities(pdf_path: str = FILES["STUDENT_GUIDE"]) -> list[dict]:
    """
    Extracts text chunks from the Student Guide PDF.
    Applies a fixed PAGE_OFFSET to align PDF numbering with InDesign layout.
    """
    from thinkcerca_tool.config import PAGE_OFFSET

    doc = fitz.open(pdf_path)
    activities = []

    for page_idx, page in enumerate(doc):
        text = page.get_text("text")
        chunks = [c.strip() for c in text.split("\n\n") if len(c.strip()) > 60]
        for chunk in chunks:
            first_line = chunk.split("\n")[0][:80]
            true_page_num = page_idx + 1 + PAGE_OFFSET  # ← fixed offset
            activities.append(
                {
                    "page": true_page_num,
                    "heading": first_line,
                    "text": chunk,
                }
            )

    print(f"✅ Extracted {len(activities)} activities (offset +{PAGE_OFFSET})")
    return activities


# ============================================================
#  AI STANDARD MATCHING
# ============================================================
def match_standards_with_ai(
    activities: list[dict],
    standards_df: pd.DataFrame,
    model: str = MODEL_NAME,
    top_k: int = 2,
) -> pd.DataFrame:
    """
    Uses GPT to match each activity with relevant standards.
    Returns DataFrame with columns: [Page, Activity, Standard Code, Reason].
    """
    client = OpenAI(api_key=OPENAI_API_KEY)
    results = []

    standards_text = "\n".join(
        f"{row.Standard_Code}: {row.Description}" for _, row in standards_df.iterrows()
    )

    for act in tqdm(activities, desc="AI Matching"):
        prompt = f"""
        You are aligning student activities with educational standards.

        Activity (from Student Guide page {act['page']}):
        \"\"\"{act['text'][:2000]}\"\"\"

        Candidate Standards:
        {standards_text}

        Pick the {top_k} most relevant standard codes.
        Return valid JSON only:
        {{
          "activity": "{act['heading']}",
          "page": {act['page']},
          "matches": [
            {{"code": "<standard code>", "reason": "<why this matches>"}}
          ]
        }}
        """
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
            )
            content = resp.choices[0].message.content.strip()

            # --- Try parsing JSON safely ---
            try:
                data = json.loads(content)
            except json.JSONDecodeError:
                match = re.search(r"\{.*\}", content, re.DOTALL)
                if match:
                    data = json.loads(match.group(0))
                else:
                    print(f"⚠️ Could not parse GPT output for page {act['page']}")
                    continue

            for m in data.get("matches", []):
                results.append(
                    {
                        "Page": act["page"],
                        "Activity": act["heading"],
                        "Standard Code": m.get("code", "").strip(),
                        "Reason": m.get("reason", "").strip(),
                    }
                )

        except Exception as e:
            print(f"⚠️ Error on page {act['page']}: {e}")

    df = pd.DataFrame(results)
    out_csv = OUTPUT_DIR / "ai_raw_matches.csv"
    df.to_csv(out_csv, index=False)
    print(f"✅ Raw AI matches saved → {out_csv}")
    return df


# ============================================================
#  PIPELINE EXECUTION
# ============================================================
def run_ai_mapping_pipeline():
    """Runs full AI mapping flow and preserves numeric page numbers."""
    print("🔍 Loading module-specific standards...")
    standards_df = join_module_standards()

    print("📘 Extracting student-guide activities...")
    activities = extract_pdf_activities()

    print("🤖 Running OpenAI LLM matching...")
    matches_df = match_standards_with_ai(activities, standards_df)

    if matches_df.empty:
        print("⚠️ No matches returned — check AI output.")
        return

    print("🧾 Formatting final workbook...")

    # --- Add metadata ---
    matches_df["Grade"] = 8
    matches_df["Unit"] = 1
    matches_df["Module"] = 2
    matches_df["Slide URL"] = ""

    # --- Ensure Page column exists ---
    if "Page" not in matches_df.columns:
        print("📄 Rebuilding Page column from activity metadata...")
        page_map = {a["heading"]: a["page"] for a in activities}
        matches_df["Page"] = matches_df["Activity"].map(page_map).fillna(-1).astype(int)

    # --- Merge with descriptions ---
    merged = pd.merge(
        matches_df,
        standards_df[["Standard_Code", "Description"]],
        left_on="Standard Code",
        right_on="Standard_Code",
        how="left"
    ).drop(columns=["Standard_Code"])
    merged.rename(columns={"Description": "Standard Description"}, inplace=True)

    merged.rename(columns={"Reason": "Slide Summary / Extracted Text"}, inplace=True)

    # --- Column order ---
    col_order = [
        "Page", "Grade", "Unit", "Module", "Activity", "Slide URL",
        "Slide Summary / Extracted Text", "Standard Code", "Standard Description"
    ]
    merged = merged[[c for c in col_order if c in merged.columns]]

    # --- Save Excel ---
    output_path = OUTPUT_DIR / "Grade8_Unit1_Module2_Mapped_Standards_AI_Final.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="Mapped Standards")

        summary_data = {
            "Project": ["ThinkCERCA AI Standards Alignment"],
            "Grade": ["8"],
            "Unit": ["1"],
            "Module": ["2 – 'I Am the Greatest'"],
            "Date": [datetime.now().strftime("%Y-%m-%d")],
            "Total Unique Activities": [merged["Activity"].nunique()],
            "Total Unique Standards": [merged["Standard Code"].nunique()],
            "Total Activity–Standard Pairs": [len(merged)],
            "Model Used": [MODEL_NAME],
            "Notes": [
                "Each row links one Activity to one or more Standards. "
                "Page numbers are auto-detected by scanning for the module start in the Student Guide PDF."
            ],
        }
        pd.DataFrame(summary_data).T.rename(columns={0: "Details"}).to_excel(
            writer, sheet_name="Summary"
        )

    # --- Beautify workbook ---
    wb = load_workbook(output_path)
    ws = wb["Mapped Standards"]

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"
    wb.save(output_path)

    print(f"✅ Final Excel with accurate page numbers → {output_path}")
